# ui/canvas.py

from PyQt5.QtWidgets import QWidget, QInputDialog, QFileDialog, QMessageBox
from PyQt5.QtGui import QPainter, QPen
from PyQt5.QtCore import Qt, QPointF, pyqtSignal
from models.shapes import Shape, LineShape, CircleShape, RectangleShape
from models.ifc_model import IFCModel
import ezdxf
import geopandas as gpd
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from shapely.geometry import shape as shapely_shape, mapping
from shapely.ops import unary_union
from network.logger import log_info, log_error

class Canvas(QWidget):
    position_changed = pyqtSignal(QPointF)
    shape_selected = pyqtSignal(object)

    def __init__(self):
        super().__init__()
        self.shapes = []
        self.current_shape = None
        self.mode = 'select'
        self.scale = 1.0
        self.offset = QPointF(0, 0)
        self.layers = {'Default': {'color': Qt.black, 'shapes': [], 'hatch': Qt.NoBrush}}
        self.current_layer = 'Default'
        self.selected_shape = None
        self.init_ui()

    def init_ui(self):
        self.setMouseTracking(True)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.scale(self.scale, self.scale)
        painter.translate(self.offset)
        self.draw_grid(painter)
        for layer_name, layer in self.layers.items():
            for shape in layer['shapes']:
                self.draw_shape(painter, shape, layer)
        if self.current_shape:
            pen = QPen(Qt.red, 2 / self.scale, Qt.DashLine)
            painter.setPen(pen)
            self.current_shape.draw(painter)
        if self.selected_shape:
            pen = QPen(Qt.blue, 2 / self.scale, Qt.DashLine)
            painter.setPen(pen)
            self.selected_shape.draw(painter)

    def draw_shape(self, painter, shape, layer):
        pen = QPen(layer['color'], 2 / self.scale)
        painter.setPen(pen)
        if 'hatch' in layer:
            brush = Qt.NoBrush if layer['hatch'] == Qt.NoBrush else Qt.BrushStyle(layer['hatch'])
            painter.setBrush(brush)
        else:
            painter.setBrush(Qt.NoBrush)
        shape.draw(painter)

    def mousePressEvent(self, event):
        pos = self.map_to_scene(event.pos())
        pos = self.snap_to_grid(pos)
        if event.button() == Qt.LeftButton:
            if self.mode == 'line':
                self.current_shape = LineShape(pos)
            elif self.mode == 'circle':
                self.current_shape = CircleShape(pos)
            elif self.mode == 'rectangle':
                self.current_shape = RectangleShape(pos)
            elif self.mode == 'select':
                self.selected_shape = self.get_shape_at_position(pos)
                self.shape_selected.emit(self.selected_shape)
                self.update()
            # 기타 모드 처리
            self.update()

    def mouseMoveEvent(self, event):
        pos = self.map_to_scene(event.pos())
        pos = self.snap_to_grid(pos)
        self.position_changed.emit(pos)
        if self.current_shape:
            self.current_shape.update(pos)
            self.update()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton and self.current_shape:
            self.layers[self.current_layer]['shapes'].append(self.current_shape)
            self.current_shape = None
            self.update()

    def map_to_scene(self, pos):
        return (pos / self.scale) - self.offset

    def snap_to_grid(self, pos):
        grid_size = 20
        x = round(pos.x() / grid_size) * grid_size
        y = round(pos.y() / grid_size) * grid_size
        return QPointF(x, y)

    def draw_grid(self, painter):
        pen = QPen(Qt.lightGray, 1 / self.scale)
        painter.setPen(pen)
        grid_size = 20
        left = int(-self.offset.x() / self.scale) - 1000
        top = int(-self.offset.y() / self.scale) - 1000
        right = int(left + self.width() / self.scale) + 2000
        bottom = int(top + self.height() / self.scale) + 2000
        for x in range(left, right, grid_size):
            painter.drawLine(x, top, x, bottom)
        for y in range(top, bottom, grid_size):
            painter.drawLine(left, y, right, y)

    def new_file(self):
        self.shapes.clear()
        self.layers = {'Default': {'color': Qt.black, 'shapes': [], 'hatch': Qt.NoBrush}}
        self.current_layer = 'Default'
        self.update()

    def open_dxf(self, filename):
        try:
            doc = ezdxf.readfile(filename)
            msp = doc.modelspace()
            self.shapes.clear()
            for entity in msp:
                if entity.dxftype() == 'LINE':
                    start = entity.dxf.start
                    end = entity.dxf.end
                    shape = LineShape(QPointF(start[0], start[1]), QPointF(end[0], end[1]))
                    self.shapes.append(shape)
                elif entity.dxftype() == 'CIRCLE':
                    center = entity.dxf.center
                    radius = entity.dxf.radius
                    shape = CircleShape(QPointF(center[0], center[1]), radius)
                    self.shapes.append(shape)
                # 기타 엔티티 처리
            self.update()
            log_info(f'DXF 파일을 열었습니다: {filename}')
        except Exception as e:
            log_error(f'파일 열기 중 오류 발생: {e}')

    def save_dxf(self, filename):
        try:
            doc = ezdxf.new(dxfversion='R2010')
            msp = doc.modelspace()
            for layer_name, layer_data in self.layers.items():
                doc.layers.add(name=layer_name)
                for shape in layer_data['shapes']:
                    shape.add_to_dxf(msp, layer_name)
            doc.saveas(filename)
            print(f'DXF 파일로 저장되었습니다: {filename}')
            log_info(f'DXF 파일로 저장되었습니다: {filename}')
        except Exception as e:
            log_error(f'파일 저장 중 오류 발생: {e}')

    def open_ifc(self, filename):
        self.ifc_model = IFCModel(filename)
        elements = self.ifc_model.get_elements()
        for element in elements:
            # 각 요소의 형상을 가져와서 캔버스에 그리기
            # 형상 데이터를 파싱하여 도형 객체로 변환
            pass
        self.update()

    def export_to_shapefile(self, filename):
        from models.shapefile_manager import ShapefileManager
        sf_manager = ShapefileManager(filename)
        # 필드 생성
        fields = [('Layer', 'C', 50)]
        if self.shapes:
            sample_shape = self.shapes[0]
            for attr_name in sample_shape.attributes.keys():
                fields.append((attr_name, 'C', 50))
            sf_manager.create_fields(fields)
            # 도형과 속성 추가
            for shape in self.shapes:
                geometry = shape.to_geometry()
                attributes = [shape.layer_name] + list(shape.attributes.values())
                sf_manager.add_record(geometry, attributes)
            sf_manager.save()

    def perform_parcel_analysis(self):
        # 연속지적도 파일 선택
        cadastral_file, _ = QFileDialog.getOpenFileName(self, '연속지적도 파일 열기', '', 'Shapefile (*.shp)')
        if not cadastral_file:
            return

        # 구획선 파일 선택
        boundary_file, _ = QFileDialog.getOpenFileName(self, '구획선 파일 열기', '', 'Shapefile (*.shp)')
        if not boundary_file:
            return

        # 지정된 필드명 입력
        field_name, ok = QInputDialog.getText(self, '필드명 입력', '속성값을 구분할 필드명을 입력하세요:')
        if not ok or not field_name:
            return

        try:
            # 지오판다스를 사용하여 Shapefile 불러오기
            cadastral_gdf = gpd.read_file(cadastral_file, encoding='utf-8')
            boundary_gdf = gpd.read_file(boundary_file, encoding='utf-8')

            # 좌표계 통일
            cadastral_gdf = cadastral_gdf.to_crs(boundary_gdf.crs)

            # 교차 분석 수행
            intersection = gpd.overlay(cadastral_gdf, boundary_gdf, how='intersection')

            # 면적 계산
            intersection['area'] = intersection.geometry.area

            # 보상금액 계산 (면적 × 공시지가)
            if '공시지가' in intersection.columns:
                intersection['compensation'] = intersection['area'] * intersection['공시지가']
            else:
                QMessageBox.warning(self, '경고', '공시지가 필드가 없습니다.')
                intersection['compensation'] = 0

            # 통계 산출
            stats = intersection.groupby(field_name).agg({
                'area': 'sum',
                'geometry': 'count',
                '공시지가': 'first',
                'compensation': 'sum'
            }).rename(columns={'geometry': '필지 수', '공시지가': '공시지가', 'area': '면적 합계', 'compensation': '보상금액'})

            # 엑셀로 내보내기
            save_file, _ = QFileDialog.getSaveFileName(self, '엑셀 파일로 저장', '', 'Excel Files (*.xlsx)')
            if save_file:
                self.export_to_excel(stats, intersection, field_name, save_file)
                QMessageBox.information(self, '완료', '엑셀 파일로 저장되었습니다.')

        except Exception as e:
            QMessageBox.critical(self, '오류', f'파셀 분석 중 오류가 발생했습니다:\n{e}')

    def export_to_excel(self, stats_df, detail_df, field_name, filename):
        wb = Workbook()
        ws_stats = wb.active
        ws_stats.title = '통계'

        # 통계 시트 작성
        ws_stats.append([field_name, '면적 합계', '필지 수', '공시지가', '보상금액'])
        for idx, row in stats_df.iterrows():
            ws_stats.append([idx, row['면적 합계'], row['필지 수'], row['공시지가'], row['보상금액']])

        # 테이블에 선 표시
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws_stats.iter_rows(min_row=1, max_row=ws_stats.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border

        # 상세 시트 작성
        ws_detail = wb.create_sheet('상세')
        ws_detail.append(detail_df.columns.tolist())
        for idx, row in detail_df.iterrows():
            ws_detail.append(row.tolist())

        # 상세 시트 테이블에 선 표시
        for row in ws_detail.iter_rows(min_row=1, max_row=ws_detail.max_row, min_col=1, max_col=ws_detail.max_column):
            for cell in row:
                cell.border = thin_border

        wb.save(filename)

    def get_shape_at_position(self, pos):
        for layer in self.layers.values():
            for shape in layer['shapes']:
                if shape.contains(pos):
                    return shape
        return None
